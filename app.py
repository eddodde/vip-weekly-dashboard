# -*- coding: utf-8 -*-
"""
VIP 주간 실적 대시보드 (LF CRM/VIP)
주간회의 엑셀 'Summary' 시트 2.실적 양식을 그대로 재현 — 수기 입력만 자동화.
표 구조: 구분(지표/채널/카테고리) × [2026년 | 전년비 | 2025년]
- 시드: data/perf_long.csv (convert.ps1 산출, tidy long)
- 사이드바에서 원본(또는 update.ps1이 만든 CSV) 업로드 시 즉시 재계산
"""
import io
import re
import calendar
import datetime
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go

st.set_page_config(page_title="VIP 주간 실적", page_icon="📊", layout="wide")

SEED_CSV = "data/perf_long.csv"

# ----------------------------------------------------------------------------- parsing
def _year_of(x):
    if x is None:
        return None
    m = re.match(r"\s*(20\d{2})", str(x))
    if m and 2000 <= int(m.group(1)) <= 2100:
        return int(m.group(1))
    return None


def _is_num(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def _grain(label):
    if label is None:
        return "unknown"
    s = str(label)
    if "/" in s:
        return "day"
    if re.search(r"\d+\D+\d+", s):
        return "week"
    if re.search(r"\d+", s):
        return "month"
    return "unknown"


def _sort_key(grain, year, label):
    s = str(label)
    if grain == "day":
        m = re.search(r"(\d{1,2})\s*/\s*(\d{1,2})", s)
        if m:
            return f"{year}{int(m.group(1)):02d}{int(m.group(2)):02d}"
    elif grain == "week":
        m = re.search(r"(\d{1,2})\D+(\d{1,2})", s)
        if m:
            return f"{year}{int(m.group(1)):02d}{int(m.group(2)):02d}"
    elif grain == "month":
        m = re.search(r"(\d{1,2})", s)
        if m:
            return f"{year}{int(m.group(1)):02d}00"
    return f"{year}9999"


def _norm_seg(x):
    if x is None:
        return ""
    s = str(x).strip()
    if s in ("-", ""):
        return ""
    if "TOTAL" in s.upper():
        return "TOTAL"
    return s


def parse_workbook(name, data):
    """Parse one BI-export xlsx (bytes) into tidy records. Mirrors convert.ps1."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not grid:
        return []
    nrows = len(grid)
    ncols = max(len(r) for r in grid)

    def cell(r, c):
        row = grid[r - 1] if r - 1 < nrows else []
        return row[c - 1] if c - 1 < len(row) else None

    val_start = 0
    for c in range(1, ncols + 1):
        if _year_of(cell(1, c)) is not None:
            val_start = c
            break
    if val_start == 0:
        return []

    year_of, label_of = {}, {}
    cur = None
    for c in range(val_start, ncols + 1):
        y = _year_of(cell(1, c))
        if y is not None:
            cur = y
        year_of[c] = cur
        label_of[c] = cell(3, c)

    grain = "unknown"
    for c in range(val_start, ncols + 1):
        if label_of[c] is not None:
            grain = _grain(label_of[c])
            break
    perspective = "product" if (val_start - 1) >= 7 else "overall"

    out = []
    cur_metric, cur_seg1 = None, None
    for r in range(2, nrows + 1):
        if not any(_is_num(cell(r, c)) for c in range(val_start, ncols + 1)):
            continue
        m = cell(r, 1)
        if m is not None:
            cur_metric = str(m).strip()
        if cur_metric is None:
            continue
        if perspective == "product":
            s1 = cell(r, 2)
            if s1 is not None:
                cur_seg1 = _norm_seg(s1)
            seg1 = cur_seg1 or ""
            seg2 = _norm_seg(cell(r, 3))
        else:
            seg1 = _norm_seg(cell(r, 2))
            seg2 = ""
        for c in range(val_start, ncols + 1):
            val = cell(r, c)
            if not _is_num(val):
                continue
            yr, lab = year_of[c], label_of[c]
            if yr is None or lab is None:
                continue
            labs = str(lab).strip()
            out.append(dict(grain=grain, perspective=perspective, year=yr,
                            period=labs, period_sort=_sort_key(grain, yr, labs),
                            seg1=seg1, seg2=seg2, metric=cur_metric,
                            value=float(val), source=name))
    return out


TIDY_COLS = ["grain", "perspective", "year", "period", "period_sort",
             "seg1", "seg2", "metric", "value", "source"]
DEDUP_KEY = ["grain", "perspective", "year", "period", "seg1", "seg2", "metric"]

# 유효회원수는 55~59k로 안정적인 스냅샷 지표. 원본(일·주·월 전 export)의 2025-09말~10월
# 구간은 이 값이 ~112k로 배증하는 데이터 손상이 있고, 같은 구간의 거래액/DAU도 함께 부풀려짐.
# → 유효회원수(TOTAL)가 연중앙값의 1.5배를 넘는 (year,period)를 손상으로 보고 전 지표에서 제외.
EFF_METRIC = "유효회원수"
# 월별은 손상 많은 월별 export 대신 깨끗한 일자별에서 집계(레벨=일평균, 비율=구성요소로 재계산).
MONTHLY_LEVEL = ["일평균거래액", "일평균고객수", "DAU", "유효회원수", "유입율"]


def _corrupt_keys(df):
    """손상 (grain,year,period): 유효회원수 또는 DAU(TOTAL)가 연중앙값의 1.5배를 넘는 구간.
    유효회원수는 55~59k 고정이라 배증이 확실한 신호, DAU는 손상 구간 경계 주차 보완용."""
    ck = set()
    for met in (EFF_METRIC, "DAU"):
        s = df[(df.perspective == "overall") & (df.metric == met) & (df.seg1 == "TOTAL")].copy()
        if s.empty:
            continue
        s["med"] = s.groupby(["grain", "year"])["value"].transform("median")
        bad = s[s["value"] > s["med"] * 1.5]
        ck |= set(zip(bad.grain, bad.year, bad.period))
    return ck


def _drop_corrupt(df):
    ck = _corrupt_keys(df)
    if not ck:
        return df, ck
    keys = zip(df["grain"], df["year"], df["period"])
    keep = [k not in ck for k in keys]
    return df[keep].copy(), ck


def derive_monthly(df):
    """깨끗한 일자별(overall)에서 월별을 집계. 손상일 제외 후 유효일<20인 월은 생략(→ '—')."""
    day = df[(df.grain == "day") & (df.perspective == "overall")].copy()
    if day.empty:
        return None
    day["ym"] = day["period_sort"].astype(str).str[:6]
    rows = []
    for (yr, ym), g in day.groupby(["year", "ym"]):
        if g[g.metric == "DAU"]["value"].count() < 20:
            continue
        mo = int(ym[4:6])
        vals = {m: g[g.metric == m]["value"].mean() for m in MONTHLY_LEVEL}
        cust, dau, sales = vals.get("일평균고객수"), vals.get("DAU"), vals.get("일평균거래액")
        vals["일평균객단가"] = sales / cust if cust else None
        vals["CR"] = cust / dau if dau else None
        for m, v in vals.items():
            if v is None or pd.isna(v):
                continue
            rows.append(dict(grain="month", perspective="overall", year=int(yr), period=f"{mo}월",
                             period_sort=f"{yr}{mo:02d}00", seg1="TOTAL", seg2="",
                             metric=m, value=float(v), source="derived_from_daily"))
    return pd.DataFrame(rows) if rows else None


def finalize(df):
    if df is None or df.empty:
        return df
    df = df.copy()
    df["year"] = df["year"].astype(int)
    # period_sort 타입 통일(str, 고정폭 8자리라 사전순=시간순).
    # CSV 시드는 int로 읽히고 업로드 파서는 str을 만들어, 병합 시 혼합되면 sort_values가 TypeError.
    df["period_sort"] = df["period_sort"].astype(str)
    for c in ("seg1", "seg2"):
        df[c] = df[c].fillna("")
    ov = df["perspective"] == "overall"
    df.loc[ov & (df["seg1"] == ""), "seg1"] = "TOTAL"
    df = df.drop_duplicates(DEDUP_KEY, keep="last").reset_index(drop=True)
    df, _ = _drop_corrupt(df)
    dm = derive_monthly(df)
    if dm is not None and not dm.empty:
        df = df[~((df.grain == "month") & (df.perspective == "overall"))]
        df = pd.concat([df, dm], ignore_index=True)
    return df.reset_index(drop=True)


def _read_tidy_csv(name, data):
    df = pd.read_csv(io.BytesIO(data), encoding="utf-8-sig")
    need = [c for c in ("grain", "perspective", "year", "period", "metric", "value") if c not in df.columns]
    if need:
        raise ValueError(f"tidy CSV 컬럼 누락: {need}")
    for c in TIDY_COLS:
        if c not in df.columns:
            df[c] = ""
    return df[TIDY_COLS]


def parse_uploads(files):
    frames, recs, drm = [], [], []
    for f in files:
        low = f.name.lower()
        try:
            if low.endswith(".csv"):
                frames.append(_read_tidy_csv(f.name, f.getvalue()))
            else:
                recs.extend(parse_workbook(f.name, f.getvalue()))
        except Exception as e:  # noqa
            if "not a zip" in str(e).lower() or "badzip" in str(e).lower():
                drm.append(f.name)
            else:
                st.sidebar.error(f"파싱 실패: {f.name} — {e}")
    if drm:
        st.sidebar.warning(
            "🔒 DRM 보호(SCDSA) 파일이라 앱에서 직접 못 엽니다: " + ", ".join(drm)
            + "\n\n로컬에서 `update.ps1` 실행 → `data/perf_long.csv` 를 올려주세요.")
    if recs:
        frames.append(pd.DataFrame(recs))
    if not frames:
        return None
    return finalize(pd.concat(frames, ignore_index=True))


@st.cache_data(show_spinner=False)
def load_seed():
    try:
        return finalize(pd.read_csv(SEED_CSV, encoding="utf-8-sig"))
    except Exception:
        return pd.DataFrame(columns=TIDY_COLS)


# ----------------------------------------------------------------------------- formatting
def fmt(metric, v):
    if v is None or pd.isna(v):
        return "-"
    if metric in ("유입율", "CR", "상품CR"):
        return f"{v*100:.1f}%"
    if metric == "일평균거래액":       # 엑셀과 동일: 백만원 정수 (558 = 5.58억)
        return f"{v/1e6:,.0f}"
    if metric in ("일평균객단가", "일평균고객수", "DAU", "유효회원수", "상품UV"):
        return f"{v:,.0f}"
    return f"{v:,.1f}"


def fmt_delta(metric, v):
    if v is None or pd.isna(v):
        return "-"
    if metric == "일평균거래액":
        return f"{v/1e6:+,.0f}"
    if metric in ("유입율", "CR", "상품CR"):
        return f"{v*100:+.1f}%p"
    return f"{v:+,.0f}"


def yoy(cur, prev):
    if prev in (None, 0) or pd.isna(prev) or cur is None or pd.isna(cur):
        return None
    return cur / prev - 1.0


def yoy_str(r):
    return "—" if r is None or pd.isna(r) else f"{r*100:+.1f}%"


def yoy_disp(r):
    """엑셀 표기: 음수=빨강 △X.X%, 양수=검정 X.X%, 결측=—. 반환 (텍스트, css)."""
    if r is None or pd.isna(r):
        return "—", ""
    if r < 0:
        return f"△{abs(r)*100:.1f}%", YOY_NEG
    return f"{r*100:.1f}%", ""


def week_pretty(lbl):
    m = re.match(r"0?(\d{1,2})\D+?(\d{1,2})", str(lbl))
    return f"{int(m.group(1))}월 {int(m.group(2))}주" if m else str(lbl)


def month_pretty(lbl):
    m = re.match(r"0?(\d{1,2})", str(lbl))
    return f"{int(m.group(1))}월" if m else str(lbl)


# ----------------------------------------------------------------------------- data load
if "df" not in st.session_state:
    st.session_state.df = load_seed()

st.sidebar.title("📊 VIP 주간 실적")
st.sidebar.caption("주간회의 'Summary' 시트 2.실적 양식")

snap_slot = st.sidebar.container()   # 이번 주 스냅샷 자리(최상단) — 내용은 데이터 로드 후 채움
st.sidebar.markdown("---")

st.sidebar.markdown(
    "#### 🧭 바로가기\n"
    "**실적 흐름**\n"
    "- [1) 거래액 트렌드](#s1)\n"
    "- [2) 월별](#s2)\n"
    "- [3) 주차별](#s3)\n"
    "- [4) 주차별·채널별](#s4)\n"
    "- [5) 행사별](#s_ev)\n"
    "- [6) 상품별](#s5)\n\n"
    "**진단·액션**\n"
    "- [✅ 종합 방향성 및 전망](#s6)"
)
st.sidebar.markdown("---")

ref_slot = st.sidebar.container()   # 📖 참고 · 진행 행사 (EVENTS 로드 후 채움)
st.sidebar.markdown("---")

with st.sidebar.expander("🔄 데이터 업데이트", expanded=False):
    st.markdown(
        "**주간 갱신**: 로컬에서 `update.ps1` 실행 → `data/perf_long.csv` 생성 → 아래 업로드.\n\n"
        "원본 BI export는 **DRM(SCDSA)** 이라 앱에서 직접 못 엽니다. (비DRM xlsx는 바로 가능)")
    st.caption("여러 파일 한 번에 선택 가능. **바뀐 파일만 올려도 병합**됩니다 "
               "(전년 일자별 등 안 바뀐 데이터는 유지). 원본이 DRM이면 `update.ps1`로 만든 CSV를 올리세요.")
    ups = st.file_uploader("CSV/xlsx 여러 개", type=["csv", "xlsx"], accept_multiple_files=True)
    if ups:
        newdf = parse_uploads(ups)
        if newdf is not None and not newdf.empty:
            base = st.session_state.df
            base = base if (base is not None and not base.empty) else load_seed()
            # 기존 + 신규 병합(같은 구간은 신규가 우선, 나머지는 유지) → 재집계
            st.session_state.df = finalize(pd.concat([base, newdf], ignore_index=True))
            st.success(f"{len(ups)}개 파일 병합 · 총 {len(st.session_state.df):,}행")
    if st.button("시드 데이터로 되돌리기"):
        load_seed.clear()
        st.session_state.df = load_seed()

df = st.session_state.df
if df is None or df.empty:
    st.warning("데이터가 없습니다. 사이드바에서 CSV/xlsx를 업로드하세요.")
    st.stop()

CUR = int(df["year"].max())
PREV = CUR - 1


# ----------------------------------------------------------------------------- accessors
def V(grain, perspective, metric, seg1, seg2, year, period):
    q = df[(df.grain == grain) & (df.perspective == perspective) & (df.metric == metric)
           & (df.seg1 == seg1) & (df.seg2 == seg2) & (df.year == year) & (df.period == period)]
    return q["value"].iloc[0] if len(q) else None


def periods(grain, perspective, metric, seg1, seg2, year):
    q = df[(df.grain == grain) & (df.perspective == perspective) & (df.metric == metric)
           & (df.seg1 == seg1) & (df.seg2 == seg2) & (df.year == year)].sort_values("period_sort")
    seen, out = set(), []
    for p in q["period"]:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


# --- 일자별/월별 MTD를 위한 날짜 유틸 (daily period_sort = YYYYMMDD) ---
_DAY = df[(df.grain == "day") & (df.perspective == "overall")].copy()
if not _DAY.empty:
    _DAY["ps"] = _DAY["period_sort"].astype(str)
    _DAY["mo"] = _DAY["ps"].str[4:6].astype(int)
    _DAY["dy"] = _DAY["ps"].str[6:8].astype(int)


def last_daily_date():
    if _DAY.empty:
        return None
    y = int(_DAY["year"].max())
    ps = _DAY[_DAY.year == y]["ps"].max()
    return datetime.date(int(ps[:4]), int(ps[4:6]), int(ps[6:8]))


def dmean(metric, year, mo, max_day=None):
    """일자별(overall,TOTAL) 월내 평균. max_day 지정 시 그 날짜까지(MTD)."""
    if _DAY.empty:
        return None
    d = _DAY[(_DAY.metric == metric) & (_DAY.seg1 == "TOTAL") & (_DAY.year == year) & (_DAY.mo == mo)]
    if max_day is not None:
        d = d[d.dy <= max_day]
    return d["value"].mean() if len(d) else None


def month_value(metric, year, mo, max_day=None):
    """월 지표값(레벨=일평균, 객단가=거래액/고객수, CR=고객수/DAU). max_day=MTD."""
    if metric == "일평균객단가":
        s, c = dmean("일평균거래액", year, mo, max_day), dmean("일평균고객수", year, mo, max_day)
        return s / c if (s is not None and c) else None
    if metric == "CR":
        c, da = dmean("일평균고객수", year, mo, max_day), dmean("DAU", year, mo, max_day)
        return c / da if (c is not None and da) else None
    return dmean(metric, year, mo, max_day)


def dv(year, mo, day, metric="일평균거래액"):
    """특정 일자 값(overall, TOTAL)."""
    if _DAY.empty:
        return None
    r = _DAY[(_DAY.metric == metric) & (_DAY.seg1 == "TOTAL") & (_DAY.year == year) & (_DAY.mo == mo) & (_DAY.dy == day)]
    return r["value"].iloc[0] if len(r) else None


def range_mean(metric, year, lo, hi):
    """[lo,hi] 기간 일자별 평균."""
    vals, d = [], lo
    while d <= hi:
        v = dv(year, d.month, d.day, metric)
        if v is not None:
            vals.append(v)
        d += datetime.timedelta(days=1)
    return sum(vals) / len(vals) if vals else None


def range_metric(metric, year, lo, hi):
    """기간 지표(레벨=일평균, 객단가=거래액/고객수, CR=고객수/DAU)."""
    if metric == "일평균객단가":
        s, c = range_mean("일평균거래액", year, lo, hi), range_mean("일평균고객수", year, lo, hi)
        return s / c if (s is not None and c) else None
    if metric == "CR":
        c, da = range_mean("일평균고객수", year, lo, hi), range_mean("DAU", year, lo, hi)
        return c / da if (c is not None and da) else None
    return range_mean(metric, year, lo, hi)


# ----------------------------------------------------------------------------- events (가변탭 행사 캘린더)
EVENTS_CSV = "data/events.csv"


@st.cache_data(show_spinner=False)
def load_events():
    try:
        e = pd.read_csv(EVENTS_CSV, encoding="utf-8-sig")
        e["text"] = e["text"].astype(str)
        e["d"] = pd.to_datetime(e["date"], errors="coerce").dt.date
        e = e[e["d"].notna()].reset_index(drop=True)          # 파싱 실패행 제거
        e["major"] = e["text"].str.contains(r"전관행사|L\+DAY|슈퍼세일|앵콜|정기세일", regex=True, na=False)
        return e
    except Exception:
        return pd.DataFrame(columns=["year", "date", "text", "d", "major"])


EVENTS = load_events()


def _evname(t):
    """'[전관행사] L+DAY (7/8 ...)' → 'L+DAY' 로 축약."""
    t = re.sub(r"\[[^\]]*\]", "", str(t))
    t = re.split(r"[(\*※]", t)[0]
    return t.strip()


def _dyear(x):
    return getattr(x, "year", None)


def upcoming_major(after_date, horizon=25):
    """오늘(after_date) 이후 horizon일 내 실제 올해(CUR) 주요 행사 [(이름, 날짜)]."""
    if EVENTS.empty or after_date is None:
        return []
    hi = after_date + datetime.timedelta(days=horizon)
    d = EVENTS["d"]
    mask = (EVENTS["major"].to_numpy(dtype=bool) & (d > after_date).to_numpy(dtype=bool)
            & (d <= hi).to_numpy(dtype=bool)
            & d.map(lambda x: _dyear(x) == CUR).to_numpy(dtype=bool))   # numpy bool(버전 무관 안전)
    e = EVENTS.loc[mask].sort_values("d")
    out, seen = [], set()
    for _, r in e.iterrows():
        nm = _evname(r["text"])
        if not nm or nm in seen:
            continue
        seen.add(nm)
        out.append((nm, r["d"]))
    return out


def event_lift(pd0, span=3):
    """행사 효과: 행사기간(pd0~+span일) 일평균거래액 vs 직전 7일. pd0는 전년 일자."""
    def avg(lo, hi):
        vals = []
        d = lo
        while d <= hi:
            v = dv(d.year, d.month, d.day)
            if v is not None:
                vals.append(v)
            d += datetime.timedelta(days=1)
        return sum(vals) / len(vals) if vals else None
    ev = avg(pd0, pd0 + datetime.timedelta(days=span - 1))
    base = avg(pd0 - datetime.timedelta(days=7), pd0 - datetime.timedelta(days=1))
    return (ev / base - 1) if (ev and base) else None


def _week_majors(lo, hi):
    """[lo,hi] 기간의 주요(전관행사급) 행사명 집합."""
    if EVENTS.empty:
        return set()
    d = EVENTS["d"]
    mask = (EVENTS["major"].to_numpy(dtype=bool) & (d >= lo).to_numpy(dtype=bool) & (d <= hi).to_numpy(dtype=bool))
    return {_evname(t) for t in EVENTS.loc[mask, "text"]}


def week_label_of(d):
    """날짜 → 주차 라벨('MM월 N주차'). ISO식(그 주 목요일이 속한 달·목요일 순번)로 회사 주차와 정렬."""
    thu = d - datetime.timedelta(days=d.weekday()) + datetime.timedelta(days=3)
    n = (thu.day - 1) // 7 + 1
    return f"{thu.month:02d}월 {n}주차"


def major_event_weeks(year):
    """해당 연도 전관행사가 걸린 주차 라벨 → 그 주 전관행사명 리스트. {label: [names]}."""
    out = {}
    if EVENTS.empty:
        return out
    mask = EVENTS["major"].to_numpy(dtype=bool) & EVENTS["d"].map(lambda d: _dyear(d) == year).to_numpy(dtype=bool)
    for _, r in EVENTS.loc[mask].iterrows():
        if "전관행사" not in str(r["text"]):
            continue
        lbl = week_label_of(r["d"])
        nm = _evname(r["text"])
        out.setdefault(lbl, [])
        if nm not in out[lbl]:
            out[lbl].append(nm)
    return out


def event_fairness(ref_date):
    """해당 주의 전년비가 '행사 정렬상 공정'한지 판정. 올해 주(ref_date 포함)의 주요행사 vs
    전년 동주(−364일) 주요행사 비교. 반환 (verdict, 설명)."""
    if EVENTS.empty or ref_date is None:
        return None
    mon = ref_date - datetime.timedelta(days=ref_date.weekday())
    sun = mon + datetime.timedelta(days=6)
    cur = _week_majors(mon, sun)
    prev = _week_majors(mon - datetime.timedelta(days=364), sun - datetime.timedelta(days=364))
    if cur == prev:
        if cur:
            return ("fair", f"양년 모두 {'·'.join(sorted(cur))} 진행 → 동일 조건(공정)")
        return ("fair", "양년 모두 이 주 전관행사 없음 → 동일 조건(공정)")
    only_prev, only_cur = prev - cur, cur - prev
    if only_prev:
        return ("warn", f"작년 이 주엔 {'·'.join(sorted(only_prev))}(전관행사)가 있어 기저가 높음 "
                        f"→ 전년비가 실제보다 나쁘게 보일 수 있음")
    return ("warn", f"올해 이 주엔 {'·'.join(sorted(only_cur))} 진행(작년엔 없음) → 전년비 상방 요인")


def _parse_period(text, year):
    """행사 텍스트의 괄호 안 '(M/D 10시 ~ M/D 10시)'에서 시작·종료일 파싱."""
    m = re.search(r"(\d{1,2})\s*/\s*(\d{1,2}).*?~\s*(\d{1,2})\s*/\s*(\d{1,2})", str(text))
    if not m:
        return None
    try:
        s = datetime.date(year, int(m.group(1)), int(m.group(2)))
        e = datetime.date(year, int(m.group(3)), int(m.group(4)))
        if e < s:
            e = datetime.date(year + 1, int(m.group(3)), int(m.group(4)))   # 연말→연초 걸침
        return s, e
    except ValueError:
        return None


def event_occurrences(year, month=None, only_major=False):
    """연도(=date의 실제연도) 행사 발생목록 [(시작, 종료, 이름, 전관여부)]. 기간은 텍스트에서 파싱,
    없으면 그리드 날짜 min~max. 이름×시작일로 중복 제거."""
    if EVENTS.empty:
        return []
    ev = EVENTS[EVENTS["d"].map(lambda d: _dyear(d) == year).to_numpy(dtype=bool)]
    if only_major:
        ev = ev[ev["major"].to_numpy(dtype=bool)]
    if ev.empty:
        return []
    recs, ranged = {}, set()
    for _, r in ev.iterrows():
        nm = _evname(r["text"])
        maj = "전관행사" in str(r["text"])
        per = _parse_period(r["text"], r["d"].year)
        if per:
            ranged.add(nm)
            key = (nm, per[0])
            if key not in recs or per[1] > recs[key][1]:
                recs[key] = (per[0], per[1], nm, maj)
    # 괄호 기간이 없는 행사 → 그리드 날짜 범위로
    bare = {}
    for _, r in ev.iterrows():
        nm = _evname(r["text"])
        if nm in ranged:
            continue
        maj = "전관행사" in str(r["text"])
        if nm not in bare:
            bare[nm] = [r["d"], r["d"], maj]
        else:
            bare[nm][0] = min(bare[nm][0], r["d"])
            bare[nm][1] = max(bare[nm][1], r["d"])
            bare[nm][2] = bare[nm][2] or maj
    for nm, (s, e, maj) in bare.items():
        recs[(nm, s)] = (s, e, nm, maj)
    out = list(recs.values())
    if month:
        out = [o for o in out if o[0].month == month]
    return sorted(out, key=lambda x: x[0])


def event_prior_lift(name, near_date=None):
    """올해 행사의 '전년 같은 행사' 효과. near_date(올해 행사일) 있으면 전년 동명 행사 중
    가장 가까운(≈−364일) 것을 골라 event_lift 계산."""
    if EVENTS.empty:
        return None
    mask = (EVENTS["d"].map(lambda x: _dyear(x) == PREV).to_numpy(dtype=bool)
            & EVENTS["text"].map(lambda t: _evname(t) == name).to_numpy(dtype=bool))
    pe = EVENTS.loc[mask]
    if pe.empty:
        return None
    if near_date is not None:
        target = near_date - datetime.timedelta(days=364)
        pe = pe.assign(_diff=pe["d"].map(lambda d: abs((d - target).days))).sort_values("_diff")
    return event_lift(pe.iloc[0]["d"])


YOY_POS = "color:#1f5fbf;font-weight:600"   # 신장(+) 파랑
YOY_NEG = "color:#c0392b;font-weight:600"   # 역신장(−) 빨강

TABLE_CSS = """
<style>
/* 제목·헤더 폰트 컴팩트화(기본값이 과대) */
.block-container{padding-top:2rem !important;}
.block-container h1{font-size:1.55rem !important;font-weight:700 !important;margin:0 0 .2rem !important;}
.block-container h2{font-size:1.15rem !important;font-weight:700 !important;margin:1rem 0 .4rem !important;}
.block-container h3{font-size:1.0rem !important;font-weight:600 !important;margin:.6rem 0 .3rem !important;}
/* 본문 metric(예상마감 등) 폰트 축소 */
.block-container [data-testid="stMetricValue"]{font-size:1.4rem !important;}
.block-container [data-testid="stMetricLabel"]{font-size:.8rem !important;}
.block-container [data-testid="stMetricLabel"] p{font-size:.8rem !important;}
.block-container [data-testid="stMetricDelta"]{font-size:.8rem !important;}
.block-container [data-testid="stMetricDelta"] svg{width:.85rem;height:.85rem;}
/* 사이드바 폰트 컴팩트화 */
[data-testid="stSidebar"] h1{font-size:1.1rem !important;font-weight:700 !important;}
[data-testid="stSidebar"] h4{font-size:.85rem !important;font-weight:700 !important;margin:.5rem 0 .2rem !important;}
[data-testid="stSidebar"] p,[data-testid="stSidebar"] li,[data-testid="stSidebar"] .stMarkdown{font-size:11.5px !important;}
[data-testid="stSidebar"] [data-testid="stCaptionContainer"]{font-size:10.5px !important;}
[data-testid="stSidebar"] [data-testid="stMetricValue"]{font-size:.95rem !important;}
[data-testid="stSidebar"] [data-testid="stMetricLabel"]{font-size:10.5px !important;}
[data-testid="stSidebar"] [data-testid="stMetricDelta"]{font-size:10.5px !important;}
[data-testid="stSidebar"] [data-testid="stMetricDelta"] svg{width:11px;height:11px;}
/* 참고 메뉴: 익스팬더 헤더·위젯 라벨을 바로가기 링크와 동일 크기로 */
[data-testid="stSidebar"] summary,[data-testid="stSidebar"] summary p{font-size:11.5px !important;font-weight:600 !important;}
[data-testid="stSidebar"] label,[data-testid="stSidebar"] label p{font-size:11px !important;}
/* 셀렉트박스 선택값 + 드롭다운 옵션 폰트(baseweb) */
section[data-testid="stSidebar"] [data-baseweb="select"],
section[data-testid="stSidebar"] [data-baseweb="select"] *,
section[data-testid="stSidebar"] [data-testid="stSelectbox"] *{font-size:11.5px !important;}
[data-baseweb="popover"] *,[data-baseweb="menu"] *,[role="listbox"] *{font-size:11.5px !important;}
.sumtbl{border-collapse:collapse;font-size:12.5px;white-space:nowrap;}
.sumtbl th,.sumtbl td{border:1px solid #d9d9d9;padding:4px 8px;text-align:right;}
.sumtbl th{background:#f2f5fa;color:#222;text-align:center;font-weight:600;}
.sumtbl td.rowh,.sumtbl th.rowh{position:sticky;left:0;background:#fafafa;text-align:left;font-weight:600;z-index:1;}
.sumtbl .grp2026{background:#eaf1fb;}
.sumtbl .grpyoy{background:#fff4ec;}
.sumtbl .grp2025{background:#f4f4f4;}
.sumtbl .grpmom{background:#eef7f1;}
.sumtbl .curcol{font-weight:700;border-left:2px solid #1f5fbf;border-right:2px solid #1f5fbf;}
.sumtbl th.curcol{border-top:2px solid #1f5fbf;}
.sumtbl .curbot{border-bottom:2px solid #1f5fbf;}
.sumwrap{overflow-x:auto;border:1px solid #e6e6e6;border-radius:6px;}
.insight{background:#f0f6ff;border-left:4px solid #1f5fbf;border-radius:4px;
  padding:8px 14px 8px 16px;margin:2px 0 12px;font-size:13.5px;line-height:1.7;}
.insight ul{margin:0;padding-left:18px;}
.insight li{margin:2px 0;}
.insight .imp{color:#1f5fbf;font-weight:600;}
/* BCG 컨설팅 스타일 종합 방향성 */
.bcg-head{background:linear-gradient(90deg,#13294b,#1f3b73);color:#fff;padding:14px 20px;border-radius:8px;
  font-size:15px;font-weight:600;line-height:1.55;margin:4px 0 16px;box-shadow:0 1px 4px rgba(0,0,0,.15);}
.bcg-head .k{color:#7fb3ff;}
.bcg-card{border:1px solid #dbe3ef;border-radius:8px;overflow:hidden;background:#fff;height:100%;}
.bcg-card h4{margin:0;padding:9px 14px;font-size:13px;color:#13294b;background:#eef3fb;
  border-bottom:2px solid #1f3b73;display:flex;justify-content:space-between;}
.bcg-card h4 .no{color:#9aa7bd;font-weight:700;}
.bcg-card ul{margin:8px 0 10px;padding:0 14px 0 28px;font-size:12.7px;line-height:1.65;}
.bcg-card li{margin:5px 0;}
.bcg-card .imp{color:#1f5fbf;font-weight:600;}
</style>
"""


_NEG_RE = re.compile(r"(△\s*[\d,.]+\s*%p?)")


def _redneg(html):
    """텍스트 속 음수(△X%)를 표와 동일하게 빨강으로. 이미 style이 걸린 곳은 건드리지 않음."""
    return _NEG_RE.sub(r'<span style="color:#c0392b;font-weight:600">\1</span>', html)


def render_insight(bullets):
    if not bullets:
        return
    lis = "".join(f"<li>{_redneg(b)}</li>" for b in bullets)
    st.markdown(f'<div class="insight"><ul>{lis}</ul></div>', unsafe_allow_html=True)


def render_gray_insight(bullets):
    """회색 톤 인사이트 박스(파란 인사이트와 형식 구분용)."""
    if not bullets:
        return
    lines = "".join(f"<div>• {_redneg(b)}</div>" for b in bullets)
    st.markdown(
        "<div style='font-size:12.5px;line-height:1.65;color:#455063;background:#f7f9fc;"
        f"border-left:3px solid #c5d3e8;padding:8px 14px;border-radius:5px;margin:2px 0 8px'>{lines}</div>",
        unsafe_allow_html=True)


def render_block_table(row_labels, blocks, bold_label=None):
    """blocks = list of (group_title, group_css, [(col_label, [cell_html per row]) ...]).
    bold_label: 해당 라벨의 열(헤더)을 강조(현재 월/주차)."""
    html = ['<div class="sumwrap"><table class="sumtbl">']
    html.append('<tr><th class="rowh" rowspan="2">구분</th>')
    for title, gcss, cols in blocks:
        html.append(f'<th colspan="{len(cols)}" class="{gcss}">{title}</th>')
    html.append('</tr><tr>')
    for title, gcss, cols in blocks:
        for cl, _ in cols:
            # 현재 열 강조는 올해·전년비 블록만 (전년 2025 블록은 참조용 → 강조 안 함)
            cur = " curcol" if (bold_label and cl == bold_label and gcss != "grp2025") else ""
            html.append(f'<th class="{gcss}{cur}">{cl}</th>')
    html.append('</tr>')
    last = len(row_labels) - 1
    for i, rl in enumerate(row_labels):
        html.append(f'<tr><td class="rowh">{rl}</td>')
        for title, gcss, cols in blocks:
            for _, cells in cols:
                cell = cells[i]
                if i == last and "curcol" in cell:   # 현재 열 하단 테두리 닫기
                    cell = cell.replace("curcol", "curcol curbot", 1)
                html.append(cell)
        html.append('</tr>')
    html.append('</table></div>')
    return "".join(html)


PERF_ROWS = [("거래액", "일평균거래액"), ("고객수", "일평균고객수"), ("DAU", "DAU"),
             ("유입률", "유입율"), ("CR", "CR"), ("객단가", "일평균객단가")]


def _td(css, inner, cur, style=""):
    stattr = f' style="{style}"' if style else ""
    return f'<td class="{css}{" curcol" if cur else ""}"{stattr}>{inner}</td>'


def perf_table(grain, cur_periods, prev_periods, pretty, bold_period=None):
    """월별/주차별 실적표: rows=지표, blocks=[2026 | 전년비 | 2025]. 전년비는 cur_periods 기준."""
    b2026, byoy, b2025 = [], [], []
    for p in cur_periods:
        cur = (p == bold_period)
        b2026.append((pretty(p), [_td("grp2026", fmt(met, V(grain, "overall", met, "TOTAL", "", CUR, p)), cur)
                                  for _, met in PERF_ROWS]))
    for p in cur_periods:
        cur = (p == bold_period)
        cells = []
        for _, met in PERF_ROWS:
            txt, sty = yoy_disp(yoy(V(grain, "overall", met, "TOTAL", "", CUR, p),
                                    V(grain, "overall", met, "TOTAL", "", PREV, p)))
            cells.append(_td("grpyoy", txt, cur, sty))
        byoy.append((pretty(p), cells))
    for p in prev_periods:
        b2025.append((pretty(p), [_td("grp2025", fmt(met, V(grain, "overall", met, "TOTAL", "", PREV, p)), False)
                                  for _, met in PERF_ROWS]))
    blocks = [(f"{CUR}년", "grp2026", b2026), ("전년비", "grpyoy", byoy), (f"{PREV}년", "grp2025", b2025)]
    return render_block_table([r for r, _ in PERF_ROWS], blocks,
                              bold_label=pretty(bold_period) if bold_period else None)


CH_ROWS = [("TTL", "TOTAL"), ("직접", "직접"), ("광고", "광고"), ("EP", "EP"), ("PUSH", "PUSH"), ("제휴", "제휴")]


def channel_table(metric, wk_periods, bold_period=None):
    """채널별 표(단일 지표): rows=채널, blocks=[2026 | 전년비 | 2025], 3블록 동일 주차 라벨."""
    b2026, byoy, b2025 = [], [], []
    for p in wk_periods:
        cur = (p == bold_period)
        b2026.append((week_pretty(p), [_td("grp2026", fmt(metric, V("week", "overall", metric, s1, "", CUR, p)), cur)
                                       for _, s1 in CH_ROWS]))
    for p in wk_periods:
        cur = (p == bold_period)
        cells = []
        for _, s1 in CH_ROWS:
            txt, sty = yoy_disp(yoy(V("week", "overall", metric, s1, "", CUR, p),
                                    V("week", "overall", metric, s1, "", PREV, p)))
            cells.append(_td("grpyoy", txt, cur, sty))
        byoy.append((week_pretty(p), cells))
    for p in wk_periods:
        b2025.append((week_pretty(p), [_td("grp2025", fmt(metric, V("week", "overall", metric, s1, "", PREV, p)), False)
                                      for _, s1 in CH_ROWS]))
    blocks = [(f"{CUR}년", "grp2026", b2026), ("전년비", "grpyoy", byoy), (f"{PREV}년", "grp2025", b2025)]
    return render_block_table([r for r, _ in CH_ROWS], blocks,
                              bold_label=week_pretty(bold_period) if bold_period else None)


def monthly_table(cur_months, cutoff_day):
    """월별 실적표. cur_months=올해 표시할 월번호(마지막=당월 MTD). 2025는 전월 1~12 전체.
    전년비: 완료월=전월마감 대비, 당월=전년 동월 MTD 대비. 당월 열 강조."""
    cur_mo = cur_months[-1]

    def cur_lbl(mo):
        return f"{mo}월(~{cutoff_day}일)" if mo == cur_mo else f"{mo}월"

    b2026, byoy, b2025 = [], [], []
    for mo in cur_months:
        is_cur = (mo == cur_mo)
        maxd = cutoff_day if is_cur else None
        cells = []
        for _, met in PERF_ROWS:
            v = month_value(met, CUR, mo, maxd) if is_cur else V("month", "overall", met, "TOTAL", "", CUR, f"{mo}월")
            cells.append(_td("grp2026", fmt(met, v), is_cur))
        b2026.append((cur_lbl(mo), cells))
    for mo in cur_months:
        is_cur = (mo == cur_mo)
        maxd = cutoff_day if is_cur else None
        cells = []
        for _, met in PERF_ROWS:
            if is_cur:
                c, p = month_value(met, CUR, mo, maxd), month_value(met, PREV, mo, maxd)
            else:
                c = V("month", "overall", met, "TOTAL", "", CUR, f"{mo}월")
                p = V("month", "overall", met, "TOTAL", "", PREV, f"{mo}월")
            txt, sty = yoy_disp(yoy(c, p))
            cells.append(_td("grpyoy", txt, is_cur, sty))
        byoy.append((cur_lbl(mo), cells))
    for mo in range(1, 13):
        b2025.append((f"{mo}월", [_td("grp2025", fmt(met, V("month", "overall", met, "TOTAL", "", PREV, f"{mo}월")), False)
                                 for _, met in PERF_ROWS]))
    blocks = [(f"{CUR}년", "grp2026", b2026), ("전년비", "grpyoy", byoy), (f"{PREV}년", "grp2025", b2025)]
    return render_block_table([r for r, _ in PERF_ROWS], blocks, bold_label=cur_lbl(cur_mo))


def _mdrange(lo, hi):
    return f"{lo.month}/{lo.day}" if lo == hi else f"{lo.month}/{lo.day}~{hi.month}/{hi.day}"


def event_period_table(cs, ce, ps, pe, ms=None, me=None):
    """전관행사 '기간' 비교: 올해(cs~ce) vs 전년(ps~pe). ms/me 주면 직전월 동종 행사(전월비)도 추가."""
    b26, byoy, b25, bmv, bmr = [], [], [], [], []
    for _, met in PERF_ROWS:
        cv, pv = range_metric(met, CUR, cs, ce), range_metric(met, PREV, ps, pe)
        b26.append(_td("grp2026", fmt(met, cv), False))
        txt, sty = yoy_disp(yoy(cv, pv))
        byoy.append(_td("grpyoy", txt, False, sty))
        b25.append(_td("grp2025", fmt(met, pv), False))
        if ms:
            mv = range_metric(met, CUR, ms, me)
            mtxt, msty = yoy_disp(yoy(cv, mv))
            bmr.append(_td("grpmom", mtxt, False, msty))
            bmv.append(_td("grp2025", fmt(met, mv), False))
    clab, plab = _mdrange(cs, ce), _mdrange(ps, pe)
    blocks = [(f"{CUR}년", "grp2026", [(clab, b26)])]
    if ms:
        blocks.append(("전월비", "grpmom", [(clab, bmr)]))
    blocks.append(("전년비", "grpyoy", [(clab, byoy)]))
    if ms:
        blocks.append(("전월", "grp2025", [(_mdrange(ms, me), bmv)]))
    blocks.append((f"{PREV}년", "grp2025", [(plab, b25)]))
    return render_block_table([r for r, _ in PERF_ROWS], blocks)


def find_prior_event(name, cur_start):
    """전년 비교 대상 행사 (start, end, name).
    ① 동명(L+DAY 등 정기 시리즈) 우선 → ② 없으면 같은 시기(±약1주) 전관행사(이름 무관)."""
    target = cur_start - datetime.timedelta(days=364)
    majors = event_occurrences(PREV, only_major=True)
    same = [o for o in majors if o[2] == name]
    if same:
        o = min(same, key=lambda o: abs((o[0] - target).days))
        if abs((o[0] - target).days) <= 21:
            return o[0], o[1], o[2]
    if majors:                             # 동명 없음 → 같은 시기 전관행사로 대조
        o = min(majors, key=lambda o: abs((o[0] - target).days))
        if abs((o[0] - target).days) <= 10:
            return o[0], o[1], o[2]
    return None


def find_prev_month_event(name, cur_start):
    """같은 해 직전 회차의 동명 행사(L+DAY 등 매월 정기 시리즈) (start, end). ~70일 이내만(결측월 방지)."""
    occ = [o for o in event_occurrences(CUR, only_major=True) if o[2] == name and o[0] < cur_start]
    if not occ:
        return None
    o = max(occ, key=lambda x: x[0])                 # 가장 최근(직전 회차)
    if (cur_start - o[0]).days > 70:
        return None
    return o[0], o[1]


def insight_event(cs, ce, ps, pe, label, ms=None, me=None):
    g = lambda m: yoy(range_metric(m, CUR, cs, ce), range_metric(m, PREV, ps, pe))
    b = _insight_sales(g, label, event=True)
    if ms:   # 직전월 동종 행사 대비(전월비) 요약 한 줄
        mm = lambda k: yoy(range_metric(k, CUR, cs, ce), range_metric(k, CUR, ms, me))
        s, a, c, dau = mm("일평균거래액"), mm("일평균객단가"), mm("일평균고객수"), mm("DAU")
        if s is not None:
            comp = {k: v for k, v in {"객단가": a, "고객수": c, "방문(DAU)": dau}.items() if v is not None}
            lead = max(comp, key=comp.get) if comp else None
            tail = f" — <b>{lead}({_pct(comp[lead])})</b>가 견인" if lead else ""
            warn = (f", 단 고객수 {_pct(c)}로 방문 회복은 미미"
                    if lead == "객단가" and c is not None and c < (a or 0) else "")
            b.append(f'<span style="color:#2e7d5b">직전월 동종행사比 거래액 <b>{_pct(s)}</b>{tail}{warn}</span>')
    return b


CATS_ORDER = ["골프", "남성", "여성", "슈즈", "잡화", "스포츠", "명품", "아웃도어", "리빙", "뷰티", "키즈"]
YEONG = ["e-영업1", "e-영업2", "e-영업3", "e-영업4"]
OWN = ["e-영업1", "e-영업2"]        # 자사 브랜드(직매입) — 액션 우선
PARTNER = ["e-영업3", "e-영업4"]    # 입점(수수료 구조)
WD = ["월", "화", "수", "목", "금", "토", "일"]  # datetime weekday(Mon=0)


PROD_METRICS = [("거래액(백만)", "일평균거래액"), ("상품UV", "상품UV")]


def product_table(wk):
    """상품별 통합표: rows=영업>카테고리, cols=거래액[26년|전년비] · 상품UV[26년|전년비]. 단일 주차."""
    def cellpair(met, ye, ca, bold):
        c = V("week", "product", met, ye, ca, CUR, wk)
        p = V("week", "product", met, ye, ca, PREV, wk)
        v = fmt(met, c)
        txt, sty = yoy_disp(yoy(c, p))
        if bold:
            v, txt = f"<b>{v}</b>", f"<b>{txt}</b>"
        return (f'<td class="grp2026">{v}</td>', f'<td class="grpyoy" style="{sty}">{txt}</td>')

    body = []
    for ye in YEONG:
        cells = []
        for _, met in PROD_METRICS:
            cells += list(cellpair(met, ye, "TOTAL", True))
        body.append((f"<b>{ye}</b>", cells))
        cats = [ca for ca in CATS_ORDER
                if V("week", "product", "일평균거래액", ye, ca, CUR, wk) is not None
                or V("week", "product", "일평균거래액", ye, ca, PREV, wk) is not None]
        for ca in cats:
            cells = []
            for _, met in PROD_METRICS:
                cells += list(cellpair(met, ye, ca, False))
            body.append(("&nbsp;&nbsp;" + ca, cells))
    html = ['<div class="sumwrap"><table class="sumtbl">']
    html.append('<tr><th class="rowh" rowspan="2">구분</th>'
                '<th colspan="2">거래액(백만)</th><th colspan="2">상품UV</th></tr>')
    html.append(f'<tr><th class="grp2026">{CUR}년</th><th class="grpyoy">전년비</th>'
                f'<th class="grp2026">{CUR}년</th><th class="grpyoy">전년비</th></tr>')
    for rl, cells in body:
        html.append(f'<tr><td class="rowh">{rl}</td>' + "".join(cells) + '</tr>')
    html.append('</table></div>')
    return "".join(html)


def product_table_one(wk, ye):
    """단일 영업의 카테고리 표(탭용): 소계 + 카테고리. cols=거래액[26|전년비]·상품UV[26|전년비]."""
    def cellpair(met, ca, bold):
        c = V("week", "product", met, ye, ca, CUR, wk)
        p = V("week", "product", met, ye, ca, PREV, wk)
        v = fmt(met, c)
        txt, sty = yoy_disp(yoy(c, p))
        if bold:
            v, txt = f"<b>{v}</b>", f"<b>{txt}</b>"
        return (f'<td class="grp2026">{v}</td>', f'<td class="grpyoy" style="{sty}">{txt}</td>')

    body = []
    cells = []
    for _, met in PROD_METRICS:
        cells += list(cellpair(met, "TOTAL", True))
    body.append((f"<b>{ye} 소계</b>", cells))
    cats = [ca for ca in CATS_ORDER
            if V("week", "product", "일평균거래액", ye, ca, CUR, wk) is not None
            or V("week", "product", "일평균거래액", ye, ca, PREV, wk) is not None]
    for ca in cats:
        cells = []
        for _, met in PROD_METRICS:
            cells += list(cellpair(met, ca, False))
        body.append((ca, cells))
    html = ['<div class="sumwrap"><table class="sumtbl">']
    html.append('<tr><th class="rowh" rowspan="2">구분</th>'
                '<th colspan="2">거래액(백만)</th><th colspan="2">상품UV</th></tr>')
    html.append(f'<tr><th class="grp2026">{CUR}년</th><th class="grpyoy">전년비</th>'
                f'<th class="grp2026">{CUR}년</th><th class="grpyoy">전년비</th></tr>')
    for rl, cells in body:
        html.append(f'<tr><td class="rowh">{rl}</td>' + "".join(cells) + '</tr>')
    html.append('</table></div>')
    return "".join(html)


# ----------------------------------------------------------------------------- charts
SALES = "일평균거래액"
BLUE_CUR, BLUE_PREV, GREY = "#1f3b73", "#9db8e0", "#9aa0a6"
CH_LINE = {"직접": "#1f3b73", "광고": "#4f7cc0", "EP": "#86a6db", "PUSH": "#c0392b"}


def _m(v):  # 원 → 백만원 (None-safe)
    return None if v is None or pd.isna(v) else v / 1e6


def _fig(title, x, series, trend=None, ypct=False):
    fig = go.Figure()
    for name, (y, color, dash) in series.items():
        fig.add_trace(go.Scatter(x=x, y=y, name=name, mode="lines+markers",
                                 line=dict(color=color, width=2, dash=dash),
                                 marker=dict(size=5), connectgaps=True))
    if trend is not None:
        xs = [i for i, v in enumerate(trend) if v is not None]
        ys = [v for v in trend if v is not None]
        if len(xs) >= 2:
            a, b = np.polyfit(xs, ys, 1)
            fig.add_trace(go.Scatter(x=x, y=[a * i + b for i in range(len(x))], name="선형",
                                     mode="lines", line=dict(color=GREY, width=1, dash="dot")))
    fig.update_layout(title=dict(text=title, font=dict(size=13)), height=300,
                      margin=dict(t=38, b=28, l=8, r=8), plot_bgcolor="white",
                      legend=dict(orientation="h", y=-0.18, font=dict(size=10)))
    fig.update_xaxes(showgrid=False, tickfont=dict(size=9))
    fig.update_yaxes(showgrid=True, gridcolor="#eee", tickfont=dict(size=9))
    if ypct:
        fig.update_yaxes(tickformat=".0%", zeroline=True, zerolinecolor="#333")
    return fig


def chart_daily(wk_labels=("", "")):
    """마지막이 일요일=2완료주, 진행중=직전 2완료주+당주(집계일까지). x축=m/d(요일). 전년은 전년동요일(−364일)."""
    last = last_daily_date()
    if last is None:
        return _fig("일자별 거래액 트렌드(전년동요일)", [], {})
    # 마지막이 일요일(완결주)이면 정확히 2주. 진행중(월~토)이면 직전 2완료주 + 당주(집계일까지).
    last_sun = last if last.weekday() == 6 else last - datetime.timedelta(days=(last.weekday() + 1) % 7)
    start = last_sun - datetime.timedelta(days=13)            # 직전 2개 완료주의 월요일
    n = (last - start).days + 1                               # 당주 진행중이면 집계일까지 연장
    days = [start + datetime.timedelta(days=i) for i in range(n)]
    x = [f"{d.month}/{d.day}({WD[d.weekday()]})" for d in days]
    y26 = [_m(V("day", "overall", SALES, "TOTAL", "", CUR, f"{d.month}/{d.day}")) for d in days]
    y25 = []
    for d in days:
        d0 = d - datetime.timedelta(days=364)
        y25.append(_m(V("day", "overall", SALES, "TOTAL", "", PREV, f"{d0.month}/{d0.day}")))
    fig = _fig("일자별 거래액 트렌드(전년동요일)", x,
               {f"{CUR}": (y26, BLUE_CUR, "solid"), f"{PREV}(동요일)": (y25, BLUE_PREV, "solid")}, trend=y26)
    # 주 경계 구분선 + 주차 라벨(월요일마다 새 주). 마지막 주는 진행중(부분).
    for gi in range(0, n, 7):
        if gi > 0:
            fig.add_vline(x=gi - 0.5, line_width=1, line_dash="dot", line_color="#bbb")
        seg = min(7, n - gi)
        lab = week_pretty(week_label_of(days[gi]))
        if gi + 7 >= n and last.weekday() != 6:   # 완결되지 않은 당주
            lab += " 진행중"
        fig.add_annotation(x=gi + (seg - 1) / 2, xref="x", yref="paper", y=1.0, yanchor="bottom",
                           text=f"<b>{lab}</b>", showarrow=False, font=dict(size=10, color="#1f3b73"))
    return fig


def chart_monthly():
    last = last_daily_date()
    cm, cd = (last.month, last.day) if last else (None, None)
    y26, y25 = [], []
    for mo in range(1, 13):
        if mo == cm:  # 당월은 MTD 일평균(일평균이라 비교 가능)
            y26.append(_m(month_value(SALES, CUR, mo, cd)))
        else:
            y26.append(_m(V("month", "overall", SALES, "TOTAL", "", CUR, f"{mo}월")))
        y25.append(_m(V("month", "overall", SALES, "TOTAL", "", PREV, f"{mo}월")))
    xlab = [f"{m}월" for m in range(1, 13)]
    fig = _fig(f"{PREV}·{CUR}년 월별 거래액 트렌드", xlab,
               {f"{CUR}": (y26, BLUE_CUR, "solid"), f"{PREV}": (y25, BLUE_PREV, "solid")})
    # 당월 예상 마감(전망) 마커 — 호버 시 근거
    fc = forecast_month(cm, cd) if cm else None
    if fc:
        ht = (f"{cm}월 예상마감 {fc['daily']/1e6:,.0f}백만<br>"
              f"근거: {cm}/1~{cd} 실적 + 잔여 {fc['rem']}일 전년 동월 실적×MTD수준"
              + (f"<br>전년 반복행사: {', '.join(fc['events'][:2])}" if fc["events"] else "") + "<extra></extra>")
        fig.add_trace(go.Scatter(x=[xlab[cm - 1]], y=[fc["daily"] / 1e6], name="예상마감",
                                 mode="markers", marker=dict(symbol="star", size=13, color="#ED7D31"),
                                 hovertemplate=ht))
    return fig


def chart_weekly(wkp):
    x = [week_pretty(p) for p in wkp]
    y26 = [_m(V("week", "overall", SALES, "TOTAL", "", CUR, p)) for p in wkp]
    y25 = [_m(V("week", "overall", SALES, "TOTAL", "", PREV, p)) for p in wkp]
    return _fig("주차별 거래액 트렌드", x,
                {f"{CUR}년": (y26, BLUE_CUR, "solid"), f"{PREV}년": (y25, BLUE_PREV, "solid")}, trend=y26)


def chart_channel_yoy(wkp):
    x = [week_pretty(p) for p in wkp]
    series = {}
    for ch, color in CH_LINE.items():
        y = [yoy(V("week", "overall", SALES, ch, "", CUR, p), V("week", "overall", SALES, ch, "", PREV, p))
             for p in wkp]
        series[ch] = (y, color, "solid")
    return _fig("주차별·채널별 거래액 전년비", x, series, ypct=True)


# ----------------------------------------------------------------------------- insights (분석형)
# 섹션 인사이트엔 '핵심 레버'만(짧게). 상세 액션(전술)은 최하단 종합 방향성에서만 노출.
DRIVER_LEVER = {"DAU": "방문(DAU) 회복이 가장 중요", "CR": "구매전환(CR) 올리기가 가장 중요",
                "객단가": "객단가 지키고 올리기가 중요"}
DRIVER_ACTION = {"DAU": "최근 미방문 VIP 자동화 문자·출석체크·개인화 추천으로 재방문 유도",
                 "CR": "장바구니·보유쿠폰·최근 본 상품 리마인드로 전환 유도",
                 "객단가": "고관여 세그 고단가 상품 소구·번들·기획전으로 객단가 상향"}


def _pct(r):
    return yoy_disp(r)[0]


def _insight_sales(g, unit_label, event=False):
    """거래액=DAU×CR×객단가 분해로 주동인·상쇄요인·핵심레버 도출(단순 현상 나열 X). g=지표→전년비.
    event=True(전관행사): VIP 기준 EP는 유입(DAU) 견인 효과가 제한적이고 CR(전환)에 효과 →
    DAU는 구조적 변수로 두고 CR을 우선 레버로 제시(방문 회복은 부수)."""
    sales = g("일평균거래액")
    comp = {"DAU": g("DAU"), "CR": g("CR"), "객단가": g("일평균객단가")}
    comp = {k: v for k, v in comp.items() if v is not None}
    if sales is None or not comp:
        return []
    neg = sales < 0
    main = (min if neg else max)(comp, key=comp.get)
    drags = [k for k in comp if (comp[k] < 0) == neg and k != main]
    defend = [k for k in comp if (comp[k] > 0) == neg]
    lead = "가 하락을 주도" if neg else "가 성장을 견인"
    also = " 동반 약세" if neg else " 동반 개선"
    b = [f"{unit_label} 거래액 <b>{_pct(sales)}</b> — <b>{main}({_pct(comp[main])})</b>{lead}"
         + (f", {drags[0]}({_pct(comp[drags[0]])}){also}" if drags else "")]
    if defend:
        b.append(("상쇄 요인" if neg else "제약 요인") + ": " + ", ".join(f"{k} {_pct(comp[k])}" for k in defend))
    if event:
        # 전관행사: DAU(유입)는 EP로 안 움직이는 구조 변수 → CR(전환)을 우선 레버, 방문은 부수.
        cr = comp.get("CR")
        if cr is not None and cr < 0:
            b.append('<span class="imp">→ 전관행사는 VIP 유입(DAU) 견인 효과가 제한적 → '
                     '<b>CR(전환) 극대화가 우선</b>(지원금 사용유도·쿠폰·리마인드), 방문(DAU)은 구조 과제로 부수 병행</span>')
        elif cr is not None:
            b.append('<span class="imp">→ CR(전환) 방어 중 — <b>전환 모멘텀 유지가 우선</b>, 방문(DAU) 회복은 부수 병행</span>')
        else:
            b.append('<span class="imp">→ <b>CR(전환) 극대화가 우선</b>, 방문(DAU)은 부수</span>')
        return b
    # (일반) 시사점 레버 = '가장 부진한(가장 낮은)' 요인. 신장 주에 이미 좋은 지표를 올리라 하지 않도록.
    worst = min(comp, key=comp.get)
    if comp[worst] < 0:
        b.append(f'<span class="imp">→ {DRIVER_LEVER.get(worst, "")}</span>')
    else:
        b.append('<span class="imp">→ 전 지표 개선, 현 상승세 유지</span>')
    return b


def _wk_sales(year, period):
    return V("week", "overall", SALES, "TOTAL", "", year, period)


def _next_prevweek(period):
    """전년 기준 다음주 라벨(계절성 참고)."""
    pw = periods("week", "overall", SALES, "TOTAL", "", PREV)
    if period in pw and pw.index(period) + 1 < len(pw):
        return pw[pw.index(period) + 1]
    return None


def resolve_weeks(latest_wk):
    """주간 데이터는 '완료주'까지 → 그 다음 주가 금주(진행중), +2가 차주.
    (라벨은 연 반복이므로 전년 주차 순서로 오프셋 계산.)"""
    pw = periods("week", "overall", SALES, "TOTAL", "", PREV)
    if not latest_wk or latest_wk not in pw:
        return None, None
    i = pw.index(latest_wk)
    geum = pw[i + 1] if i + 1 < len(pw) else None    # 금주(진행중)
    cha = pw[i + 2] if i + 2 < len(pw) else None      # 차주
    return geum, cha


def forecast_month(mo, cutoff):
    """당월 예상 마감. 잔여일은 '전년 동요일(−364)' 실적으로 채우고(반복 행사·주말 자동 반영),
    올해 MTD 수준(전년 동기比)만큼 스케일. → 행사 일자가 1~2일 달라도 요일 기준으로 정렬됨."""
    if not cutoff or _DAY.empty:
        return None
    dim = calendar.monthrange(CUR, mo)[1]
    if cutoff >= dim:
        return None
    def s(year, lo, hi):
        return sum(v for v in (dv(year, mo, d) for d in range(lo, hi + 1)) if v is not None)
    mtd26, mtd25, rem25 = s(CUR, 1, cutoff), s(PREV, 1, cutoff), s(PREV, cutoff + 1, dim)
    if mtd25 <= 0:
        return None
    # 잔여일은 전년 동월 실적(반복 행사 포함)에 올해 MTD 수준(전년비 ratio)을 반영.
    # 월 단위 집계라 행사가 1~2일 이동해도 총액 영향은 미미 → 전년비는 MTD와 일관(=ratio−1).
    ratio = mtd26 / mtd25
    proj_total = mtd26 + rem25 * ratio
    proj_daily = proj_total / dim
    ev = [nm for nm, d in upcoming_major(datetime.date(CUR, mo, cutoff), horizon=dim - cutoff)]
    return dict(daily=proj_daily, total=proj_total, dim=dim, ratio=ratio, rem=dim - cutoff,
                yoy=ratio - 1, events=ev)


def insight_trend(wk_all):
    """최신주(지난주 마감) 전년비·전주비 + 전년 동기 근거로 '금주(진행중)' 전망."""
    if not wk_all:
        return []
    period = wk_all[-1]                         # 최신 완료주 = 지난주(마감)
    geum, cha = resolve_weeks(period)           # 금주(진행중), 차주
    b = []
    yoyv = yoy(_wk_sales(CUR, period), _wk_sales(PREV, period))
    wowv = yoy(_wk_sales(CUR, period), _wk_sales(CUR, wk_all[-2])) if len(wk_all) >= 2 else None
    b.append(f"지난주 마감({week_pretty(period)}) 거래액 전년비 <b>{_pct(yoyv)}</b>"
             + (f" · 전주비 <b>{_pct(wowv)}</b>" if wowv is not None else ""))
    fr = event_fairness(last_daily_date())
    if fr:
        icon = "✅ 전년비 공정" if fr[0] == "fair" else "⚠️ 전년비 주의"
        b.append(f'<span style="color:#5b6472"><b>{icon}</b> — {fr[1]}</span>')
    if geum:
        a, c = _wk_sales(PREV, period), _wk_sales(PREV, geum)   # 전년 지난주, 전년 금주(동주)
        seas = yoy(c, a)
        if a and c and seas is not None:
            b.append(f"작년 같은 시기엔 {week_pretty(period)}→{week_pretty(geum)} "
                     f"{a/1e6:,.0f}→{c/1e6:,.0f}백만(<b>{_pct(seas)}</b>)로 {'올랐음' if seas > 0 else '내렸음'} "
                     f'→ <span class="imp">올해 <b>금주({week_pretty(geum)})</b>도 '
                     f'{"반등 기대" if seas > 0 else "약세 이어질 수 있어 방어 필요"}</span>')
    up = upcoming_major(last_daily_date(), horizon=10)
    if up:
        nm, ed = up[0]
        lift = event_prior_lift(nm, ed)
        liftxt = f" (작년 이 행사 때 거래액 직전주보다 <b>{_pct(lift)}</b>)" if lift is not None else ""
        b.append(f"<b>금주 {nm}</b>({ed.month}/{ed.day} 진행) 예정{liftxt} "
                 f'→ <span class="imp">행사 효과로 반등 유도, 사전 알림·고관여 타겟팅</span>')
    return b


def insight_perf(grain, period, unit_label):
    g = lambda m: yoy(V(grain, "overall", m, "TOTAL", "", CUR, period),
                      V(grain, "overall", m, "TOTAL", "", PREV, period))
    return _insight_sales(g, unit_label)


def insight_month(mo, cutoff, is_cur, unit_label):
    maxd = cutoff if is_cur else None
    g = lambda m: yoy(month_value(m, CUR, mo, maxd), month_value(m, PREV, mo, maxd))
    b = _insight_sales(g, unit_label)
    if is_cur and mo > 1 and b:
        mom = yoy(month_value("일평균거래액", CUR, mo, maxd), month_value("일평균거래액", CUR, mo - 1, cutoff))
        if mom is not None:
            b.insert(1, f"전월비(동일기간 MTD) 거래액 <b>{_pct(mom)}</b>")
    if is_cur:
        fc = forecast_month(mo, cutoff)
        if fc and fc["yoy"] is not None:
            evtxt = f" (잔여기간 {', '.join(fc['events'][:2])} 등 전년 반복 행사 포함)" if fc["events"] else ""
            b.insert(len(b) - 1, f"<b>{mo}월 예상 마감</b> 일평균 <b>{fc['daily']/1e6:,.0f}백만</b>"
                                 f"(전년비 {_pct(fc['yoy'])}, 월 거래액 약 {fc['total']/1e8:,.0f}억){evtxt}")
    return b


def insight_dau(period):
    """'DAU 견인 필요'(누구나 아는 얘기)를 넘어: 지속성·주도 채널·행사 무관 구조성까지."""
    rows = []
    for _, ch in CH_ROWS[1:]:
        c, p = V("week", "overall", "DAU", ch, "", CUR, period), V("week", "overall", "DAU", ch, "", PREV, period)
        if c is not None and p is not None:
            rows.append((ch, c - p, yoy(c, p)))
    if not rows:
        return []
    worst = min(rows, key=lambda r: r[1])
    wa = periods("week", "overall", "DAU", "TOTAL", "", CUR)
    cons = 0
    for pp in reversed(wa):
        r = yoy(V("week", "overall", "DAU", "TOTAL", "", CUR, pp), V("week", "overall", "DAU", "TOTAL", "", PREV, pp))
        if r is not None and r < 0:
            cons += 1
        else:
            break
    b = [f"DAU가 <b>{cons}주 연속</b> 전년보다 낮음(잠깐이 아니라 계속되는 하락) — "
         f"특히 <b>{worst[0]} 채널</b>에서 많이 빠짐({_pct(worst[2])})"]
    b.append('<span class="imp">→ 행사 때 잠깐 오는 방문이 아니라 <b>평상시 방문</b>(안 오던 VIP 다시 부르기·앱 재방문) 회복이 핵심</span>')
    return b


def insight_channel(period):
    """비중(규모)을 반영해 실질 진원지·선전 채널만 짚음(방법론은 노출하지 않음)."""
    tc = V("week", "overall", SALES, "TOTAL", "", CUR, period)
    rows = []
    for _, ch in CH_ROWS[1:]:
        c, p = V("week", "overall", SALES, ch, "", CUR, period), V("week", "overall", SALES, ch, "", PREV, period)
        if c is None or p is None:
            continue
        rows.append((ch, c - p, (c / tc if tc else 0), yoy(c, p)))
    if not rows:
        return []
    big = [r for r in rows if r[2] >= 0.05] or rows          # 규모 있는 채널 위주 해석
    sw = sorted(big, key=lambda r: r[1])
    worst = min(rows, key=lambda r: r[1])
    lift = sw[-1]
    b = [f"거래액이 가장 많이 빠진 채널은 <b>{worst[0]}</b>(비중 {worst[2]*100:.0f}%) — {worst[1]/1e6:+,.0f}백만, {_pct(worst[3])}"]
    sec = [r for r in sw if r[1] < 0 and r[0] != worst[0]]
    b.append((f"큰 채널 중 {sec[0][0]}({_pct(sec[0][3])})도 부진, " if sec else "")
             + f"<b>{lift[0]}</b>(비중 {lift[2]*100:.0f}%, {_pct(lift[3])})는 그나마 선방")
    b.append(f'<span class="imp">→ 비중 큰 {worst[0]}부터 살려야 전체가 반등</span>')
    return b


def insight_product(wk):
    """자사(영업1·2) 중심으로 부진·견조 카테고리를 짚음."""
    recs = []
    for ye in YEONG:
        for ca in CATS_ORDER:
            c = V("week", "product", SALES, ye, ca, CUR, wk)
            p = V("week", "product", SALES, ye, ca, PREV, wk)
            if c is not None and p is not None:
                recs.append((ye, ca, f"영업{ye[-1]} {ca}", c - p, yoy(c, p)))
    if not recs:
        return []
    allw = sorted(recs, key=lambda r: r[3])[:2]
    b = ["전체 하락 주도: " + ", ".join(f"<b>{r[2]}</b>({_pct(r[4])})" for r in allw)]
    own = [r for r in recs if r[0] in OWN]
    if own:
        ow = min(own, key=lambda r: r[3])
        ob = max(own, key=lambda r: r[3])
        b.append(f"자사(영업1·2): <b>{ow[2]}</b> 부진({_pct(ow[4])}) / <b>{ob[2]}</b> 견조({_pct(ob[4])})")
        b.append(f'<span class="imp">→ 자사 {ow[1]} 반등이 우선, {ob[1]} 성장세 유지</span>')
    return b


def insight_yeong(ye, wk):
    """영업 탭별 핵심 인사이트(간단·명료): 소계 전년비 + 부진/신장 카테고리 + 시사점."""
    c = V("week", "product", SALES, ye, "TOTAL", CUR, wk)
    p = V("week", "product", SALES, ye, "TOTAL", PREV, wk)
    if c is None or p is None:
        return []
    tot = yoy(c, p)
    recs = []
    for ca in CATS_ORDER:
        cc = V("week", "product", SALES, ye, ca, CUR, wk)
        pp = V("week", "product", SALES, ye, ca, PREV, wk)
        if cc is not None and pp is not None:
            recs.append((ca, cc - pp, yoy(cc, pp)))
    own = ye in OWN
    seg = "자사" if own else "입점"
    b = [f"<b>{ye}</b>({seg}) 거래액 전년비 <b>{_pct(tot)}</b>"]
    if recs:
        srt = sorted(recs, key=lambda r: r[1])
        worst, best = srt[0], srt[-1]
        parts = []
        if worst[1] < 0:
            parts.append(f"<b>{worst[0]}</b> 부진({_pct(worst[2])})")
        if best[2] is not None and best[2] > 0:
            parts.append(f"<b>{best[0]}</b> 신장({_pct(best[2])})")
        if parts:
            b.append(" · ".join(parts))
        if own and worst[1] < 0:
            tail = f", {best[0]} 성장세 유지" if (best[2] and best[2] > 0) else ""
            b.append(f'<span class="imp">→ 자사 <b>{worst[0]}</b> 반등 우선(시크릿 혜택·기획전){tail}</span>')
        elif own:
            b.append(f'<span class="imp">→ {best[0]} 호조 확대로 자사 비중 강화</span>')
        else:
            b.append(f'<span class="imp">→ 입점 채널 · <b>{worst[0]}</b> 약세 점검, 자사 우선</span>')
    return b


def final_direction(wk_all, cur_mo, cutoff):
    """진단 → 금주 액션 → 차주 정량 전망(핵심동인 회복 시나리오)."""
    if not wk_all:
        return [], [], []
    period = wk_all[-1]
    g = lambda m: yoy(V("week", "overall", m, "TOTAL", "", CUR, period),
                      V("week", "overall", m, "TOTAL", "", PREV, period))
    sales = g("일평균거래액")
    comp = {"DAU": g("DAU"), "CR": g("CR"), "객단가": g("일평균객단가")}
    comp = {k: v for k, v in comp.items() if v is not None}
    main = min(comp, key=comp.get) if comp else None
    # 자사 부진 카테고리
    orecs = []
    for ye in OWN:
        for ca in CATS_ORDER:
            c = V("week", "product", SALES, ye, ca, CUR, period)
            p = V("week", "product", SALES, ye, ca, PREV, period)
            if c is not None and p is not None:
                orecs.append((ca, c - p, yoy(c, p)))
    own_worst = min(orecs, key=lambda r: r[1])[0] if orecs else None
    geum, cha = resolve_weeks(period)                       # 금주(진행중), 차주
    gw = week_pretty(geum) if geum else "금주"
    seas = yoy(_wk_sales(PREV, geum), _wk_sales(PREV, period)) if geum else None
    last = last_daily_date()
    up = upcoming_major(last, horizon=25)                   # 실제 올해 행사
    ge = next((e for e in up if e[1] <= last + datetime.timedelta(days=7)), None)   # 금주 행사
    later = [e for e in up if not ge or e[1] > ge[1]]
    ce = later[0] if later else None                        # 이후 다음 행사

    dau, cr, aov = comp.get("DAU"), comp.get("CR"), comp.get("객단가")
    drags = [k for k in ("DAU", "CR", "객단가") if comp.get(k, 0) < 0]
    buf = [k for k in ("DAU", "CR", "객단가") if comp.get(k, 0) > 0]
    # 즉효(전술) 레버 = 전환(CR): 행사 고트래픽 위에서 단기에 움직임 / 구조 레버 = 방문(DAU)
    tactical = "CR" if (cr is not None and cr < 0) else None

    diag, now, nxt = [], [], []
    # 진단: 퍼널 어디가 새는지 + DAU 구조성
    if sales is not None:
        drag_txt = "·".join(f"{k}({_pct(comp[k])})" for k in drags) or "-"
        buf_txt = (", ".join(f"{k}({_pct(comp[k])})" for k in buf) + "만 상방 기여") if buf else ""
        diag.append(f"거래액 <b>{_pct(sales)}</b> = <b>{drag_txt} 동반 부진</b>{('; ' + buf_txt) if buf_txt else ''} "
                    f"→ 새는 곳은 <b>상단(방문)·중단(전환) 퍼널</b>, 고가치 구매층은 유지")
    diag += insight_dau(period)     # DAU 지속성·주도채널·평상시 방문
    # 금주 액션(진행중): 전환(CR)을 최우선으로
    if tactical and ge:
        now.append(f"<b>{ge[0]} 고트래픽에 전환(CR) 화력 집중</b> — {DRIVER_ACTION['CR']} "
                   f"(CR이 방문만큼 빠졌고, 방문 몰릴 때 즉시 만회 가능)")
    else:
        now.append(f"구매전환 지키기 — {DRIVER_ACTION['CR']}")
    if own_worst:
        now.append(f"자사 부진 <b>{own_worst}</b> 구매전환(시크릿 혜택·기획전) + 잘 나가는 카테고리 밀어주기")
    now.append(f"방문(DAU)은 구조적 하락 → 행사 유입객을 상시 재방문으로 묶기(앱 푸시 재동의·개인화 홈)")
    if ge:
        now.append(f"<b>금주 {ge[0]}</b>({ge[1].month}/{ge[1].day} 진행중) 사전 알림톡·고관여 타겟 집중")

    # 전망 ① 금주 절대수준: 작년 행사주(주차별 실적) 반등폭을 올해 지난주에 적용
    evnm = ge[0] if ge else "행사"
    if geum:
        lp, lg, tp = _wk_sales(PREV, period), _wk_sales(PREV, geum), _wk_sales(CUR, period)
        if lp and lg and tp:
            wowL = lg / lp - 1
            est = tp * (1 + wowL)
            nxt.append(f"금주 <b>{evnm}</b>로 절대 거래액 반등 — 올해 지난주 {tp/1e6:,.0f}백만에 작년 행사주 반등폭"
                       f"({wowL*100:+.0f}%, {lp/1e6:,.0f}→{lg/1e6:,.0f}백만) 적용 시 <b>금주 ~{est/1e6:,.0f}백만</b>")
            nxt.append(f"단 작년에도 이 주에 {evnm}가 있었기 때문에 <b>행사만으론 전년비({_pct(sales)}) 불변</b> — "
                       f"전년비를 되돌리려면 작년보다 잘 팔아야 함")
    # 전망 ② 어느 레버가 이번 주에 실제로 움직이나(핵심 인사이트)
    if tactical:
        scen = sales - cr            # 전환을 작년 수준으로 회복(CR 격차 만회)
        half = sales - cr / 2
        nxt.append(f"이번 주 <b>즉효 레버는 전환(CR {_pct(cr)})</b> — 방문(DAU)은 9주째 구조적이라 단기 반전이 어렵지만, "
                   f"전환은 {evnm} 고트래픽 위에서 전술로 당길 수 있음")
        nxt.append(f"{evnm}주 전환율을 작년 수준까지 회복하면 거래액 전년비 <b>{_pct(sales)} → {_pct(scen)}</b>"
                   f"(절반만 잡아도 {_pct(half)}) — 방문 회복(구조)은 다음 분기 과제로 병행")
    if ce:
        nxt.append(f"이후 <b>{ce[0]}</b>({ce[1].month}/{ce[1].day}) 전관행사가 이어져 반등 흐름 연장 가능")
    nxt.append("중기: 방문(DAU) 구조 회복 — 직접/앱 재방문 프로그램으로 상시 트래픽 복원")

    # 헤드라인: 어디가 새고, 이번 주 무엇을 당길지(비자명한 우선순위)
    head = ""
    if sales is not None:
        if tactical:
            scen = sales - cr
            head = (f"지난주 {_pct(sales)}는 <span class='k'>방문·전환 동반 부진</span> — 방문(DAU)은 구조적이라 오래 걸리니, "
                    f"이번 주 <span class='k'>{evnm} 고트래픽에서 '전환(CR)'을 끌어올리는 것</span>이 즉효 레버"
                    f"(작년 전환율 회복 시 <span class='k'>{_pct(sales)}→{_pct(scen)}</span>)")
        else:
            head = f"지난주 거래액 {_pct(sales)} — 방문(DAU) 구조 회복이 근본 과제"
    return head, diag, now, nxt


def render_bcg(head, diag, now, nxt):
    """BCG 컨설팅 스타일: 핵심 메시지 헤드라인 + 진단/실행/임팩트 3-카드."""
    if head:
        st.markdown(f'<div class="bcg-head">💡 {head}</div>', unsafe_allow_html=True)
    cards = [("① 진단", "지난주 마감 · Where we are", diag),
             ("② 금주 실행", "진행중 · What to do now", now),
             ("③ 전망", "금주·차주 · Impact", nxt)]
    cols = st.columns(3)
    for col, (title, sub, bullets) in zip(cols, cards):
        lis = "".join(f"<li>{_redneg(b)}</li>" for b in bullets) or "<li>-</li>"
        col.markdown(f'<div class="bcg-card"><h4><span>{title}</span><span class="no">{sub}</span></h4>'
                     f'<ul>{lis}</ul></div>', unsafe_allow_html=True)


# ============================================================================= RENDER
st.markdown(TABLE_CSS, unsafe_allow_html=True)

wk_all = periods("week", "overall", "일평균거래액", "TOTAL", "", CUR)
latest_wk = wk_all[-1] if wk_all else None
st.title(f"■ {week_pretty(latest_wk) if latest_wk else ''} 마감 CRM_VIP 실적")
st.caption(f"기준연도 {CUR} · 전년 {PREV}  |  주간회의 Summary 시트 2.실적 양식 · 자동 집계 "
           f"· **모든 실적은 일평균 기준**(거래액=일평균거래액, 단위 백만원)")

# 사이드바 최상단 스냅샷(예약 슬롯 채우기) — 최신주 전년비 KPI
if latest_wk:
    def _snap(col, label, met, is_sales=False):
        c = V("week", "overall", met, "TOTAL", "", CUR, latest_wk)
        p = V("week", "overall", met, "TOTAL", "", PREV, latest_wk)
        val = "-" if c is None else (f"{c/1e6:,.0f}백만" if is_sales else fmt(met, c))
        col.metric(label, val, yoy_str(yoy(c, p)))
    with snap_slot:
        st.markdown("#### 📌 이번 주 스냅샷")
        st.caption(f"{week_pretty(latest_wk)} 마감 · **일평균** 기준 · 전년비")
        r1 = st.columns(2)
        _snap(r1[0], "거래액", "일평균거래액", is_sales=True)
        _snap(r1[1], "DAU", "DAU")
        r2 = st.columns(2)
        _snap(r2[0], "CR", "CR")
        _snap(r2[1], "객단가", "일평균객단가")
        st.caption(f"🗓 집계 기준일 **{last_daily_date()}** — 이 날짜가 최신 데이터입니다")

# 사이드바 참고 슬롯 채우기 — 진행 행사 캘린더(전년/올해)
with ref_slot:
    st.markdown("#### 📖 참고 · 진행 행사")
    with st.expander("행사 캘린더 보기", expanded=False):
        _cm = last_daily_date().month if last_daily_date() else 0   # 당월 기본
        yr = st.selectbox("연도", [PREV, CUR], index=0, format_func=lambda y: f"{y}년", key="ev_yr")
        mo = st.selectbox("월", ["전체"] + [f"{m}월" for m in range(1, 13)], index=_cm, key="ev_mo")
        only_major = st.checkbox("전관행사만", value=False, key="ev_major")
        occs = event_occurrences(yr, None if mo == "전체" else int(mo[:-1]), only_major)
        if not occs:
            st.caption("표시할 행사가 없습니다.")
        else:
            md = "\n".join(
                f"- **{s.month}/{s.day}\\~{e.month}/{e.day}** · {'★ ' if maj else ''}{nm}"
                for s, e, nm, maj in occs)
            st.markdown(md)
            st.caption("기간=시작~종료 · ★=전관행사(전사)")

wk_periods = wk_all[-5:]                       # 엑셀과 동일하게 최근 5주 고정
cutoff = last_daily_date().day if last_daily_date() else None
cur_months = [int(p[:-1]) for p in periods("month", "overall", "일평균거래액", "TOTAL", "", CUR)]
if cutoff and (not cur_months or cur_months[-1] != last_daily_date().month):
    cur_months.append(last_daily_date().month)  # 당월(MTD) 추가

# ---- 1) 거래액 트렌드 ----
st.header("1) 거래액 트렌드", anchor="s1")
render_insight(insight_trend(wk_all))
wk_lbls = (week_pretty(wk_all[-2]) if len(wk_all) >= 2 else "", week_pretty(latest_wk) if latest_wk else "")
cc = st.columns(4)
cc[0].plotly_chart(chart_daily(wk_lbls), use_container_width=True)
cc[1].plotly_chart(chart_monthly(), use_container_width=True)
cc[2].plotly_chart(chart_weekly(wk_periods), use_container_width=True)
cc[3].plotly_chart(chart_channel_yoy(wk_periods), use_container_width=True)

# ---- 2) 월별 ----
st.header("2) 월별", anchor="s2")
cur_mo = cur_months[-1] if cur_months else None
render_insight(insight_month(cur_mo, cutoff, True, f"당월({cur_mo}월~{cutoff}일 MTD)") if cur_mo else [])
st.caption(f"전년비 — 완료월: 전년 동월 **마감** 대비 / 당월({cur_mo}월~{cutoff}일): 전년 동월 **동일기간(MTD)** 대비")
st.markdown(monthly_table(cur_months, cutoff), unsafe_allow_html=True)
# 하단 참고: 예상 마감(추정치) — 작게, 표 아래에
fc = forecast_month(cur_mo, cutoff) if cur_mo else None
if fc and fc["yoy"] is not None:
    basis = (f"{cur_mo}월 1~{cutoff}일 실적 + 잔여 {fc['rem']}일은 전년 동월 같은 일자 실적에 "
             f"올해 MTD 수준(전년비 {fc['ratio']*100-100:+.0f}%)을 반영해 추정. "
             f"전년 대비 행사 컨텐츠 동일·일자만 1~2일 이동 가정(월 총액 영향 미미)"
             + (f". 잔여기간 전년 반복 행사: {', '.join(fc['events'][:3])}" if fc["events"] else "") + ".")
    st.markdown(
        f"<div style='font-size:12px;line-height:1.55;color:#5b6472;background:#f7f9fc;"
        f"border-left:3px solid #c5d3e8;padding:7px 12px;border-radius:4px;margin-top:8px'>"
        f"📌 <b>참고 · {cur_mo}월 예상 마감(추정치, 변동 가능)</b> — "
        f"일평균 <b>{fc['daily']/1e6:,.0f}백만</b>(전년비 {yoy_str(fc['yoy'])}) · "
        f"월 거래액 약 <b>{fc['total']/1e8:,.0f}억</b> · MTD(~{cutoff}일) 전년비 {yoy_str(fc['ratio']-1)}"
        f"<br><span style='color:#93a0b3'>근거: {basis}</span></div>",
        unsafe_allow_html=True)

# 하단 참고: 구조 진단(수동 갱신 — 2026.7월 심층분석 요약. 상황 변하면 이 블록 업데이트)
st.markdown(
    "<div style='font-size:12px;line-height:1.7;color:#5b4a32;background:#fdf9f1;"
    "border-left:3px solid #d9b45c;padding:8px 12px;border-radius:4px;margin-top:6px'>"
    "🔍 <b>참고 · 구조 진단 (7월 심층분석 요약)</b>"
    "<br>· <b>고객수 감소는 저단가 구매 위축이 주도</b> — 객단가 하위 브랜드 구매고객 △30~46%, 20~40만 프리미엄대만 +10%. "
    "객단가 상승은 그 반사효과(총거래액 미성장·퍼널 위축 → 프리미엄화 아님)"
    "<br>· <b>슈즈 급감(△40%)은 '25.9월 핏플랍 철수 공백</b>이 샌들 성수기에 표면화된 것"
    "(전년 카테 거래액의 49%·고객의 56%, 잔여 브랜드 흡수 19%) — 7~8월 지속, 9월부터 기저 소멸"
    "<br>· <b>여름은 저단가 상품이 전환(CR)을 견인하는 시즌</b>(객단가↔CR 역상관 △0.76)이라 "
    "저단가 구색 공백이 6월부터 CR 역신장으로 전이 — 행사·앱 사용성 요인은 데이터상 무관"
    "(행사일/비행사일 하락폭 동일, 인당 조회수 전년 동일)"
    "<br>· <b>상반기 CR 우위(+5~10%)는 전년 PUSH·광고 전환 개선의 기저효과</b>로 6월 소진. "
    "직접 채널 CR 6월부터 첫 하락(△6~8%), 광고는 유입 +21% 확대에도 CR △12%(효율 저하)"
    "<br>· <b>방문 기반: 회원 수 유지, 방문율만 하락(35→32%)</b> — 감소분은 직접·PUSH 채널(습관성 방문). "
    "PUSH는 전환력 유지 중이라 도달 회복 여지"
    "</div>", unsafe_allow_html=True)

# ---- 3) 주차별 ----
st.header("3) 주차별", anchor="s3")
render_insight(insight_perf("week", latest_wk, f"최신주({week_pretty(latest_wk)})"))
st.markdown(perf_table("week", wk_periods, wk_periods, week_pretty, bold_period=latest_wk), unsafe_allow_html=True)

# ---- 4) 주차별·채널별 ----
st.header("4) 주차별·채널별", anchor="s4")
render_insight(insight_channel(latest_wk))
for tag, met in [("① 거래액", "일평균거래액"), ("② DAU", "DAU"), ("③ CR", "CR")]:
    st.subheader(tag)
    st.markdown(channel_table(met, wk_periods, bold_period=latest_wk), unsafe_allow_html=True)

# ---- 5) 행사별 (전년 동일 행사 비교) ----
st.header("5) 행사별 (전년·전월 비교)", anchor="s_ev")
_ld = last_daily_date()
# 시작된 전관행사(진행중 포함): 시작일 ≤ 집계일
_started = sorted([o for o in event_occurrences(CUR, only_major=True) if _ld and o[0] <= _ld],
                  key=lambda o: o[0])


def _evsel_label(o):
    if o[1] > _ld:                                   # 진행중
        return f"{o[2]} · 진행중 {(_ld - o[0]).days + 1}일차 ({_mdrange(o[0], _ld)})"
    return f"{o[2]} · 종료 ({_mdrange(o[0], o[1])})"


if _started:
    sel_ev = st.selectbox("전관행사 선택", _started[::-1], index=0, key="ev_week", format_func=_evsel_label)
    cs, ce, nm = sel_ev[0], sel_ev[1], sel_ev[2]
    inprog = ce > _ld
    ce_eff = min(ce, _ld)                            # 올해 유효 종료일(진행중이면 집계일)
    elapsed = (ce_eff - cs).days + 1
    prev = find_prior_event(nm, cs)
    if prev:
        ps, pe, pn = prev
        pe_eff = (ps + datetime.timedelta(days=elapsed - 1)) if inprog else pe   # 전년 같은 경과일
        unit = f"{nm} {elapsed}일차까지" if inprog else f"{nm} 기간"
        pm = find_prev_month_event(nm, cs)          # 직전월 동종 행사(L+DAY 등)
        ms = me = None
        if pm:
            ms = pm[0]
            me = ms + datetime.timedelta(days=(ce_eff - cs).days)   # 같은 경과일수로 정렬
        render_insight(insight_event(cs, ce_eff, ps, pe_eff, unit, ms, me))
        _c, _p = _mdrange(cs, ce_eff).replace("~", "\\~"), _mdrange(ps, pe_eff).replace("~", "\\~")
        _stat = f"진행중 {elapsed}일차까지 · 행사 경과일 정렬" if inprog else "행사 종료 · 전체 기간"
        if pn == nm:
            st.caption(f"올해 **{nm}** {_c} ↔ 전년 **{nm}** {_p} · {_stat}")
        else:
            st.caption(f"올해 **{nm}** {_c} ↔ 전년 동기 전관행사 **{pn}** {_p} · {_stat}(행사명 다름)")
        if pm:
            _mm = _mdrange(ms, me).replace("~", "\\~")
            st.caption(f"↕ **전월비**: 직전월 **{nm}** {_mm} 대비 (매월 정기 행사)")
        st.markdown(event_period_table(cs, ce_eff, ps, pe_eff, ms, me), unsafe_allow_html=True)
    else:
        st.caption("⚠️ 전년 같은 시기에 전관행사가 없어 비교 불가")
else:
    st.caption("아직 시작된 전관행사가 없습니다.")

# ---- 6) 상품별 ----
if not df[df.perspective == "product"].empty:
    st.header("6) 상품별 (e-영업 × 카테고리)", anchor="s5")
    pwk_all = periods("week", "product", "일평균거래액", "e-영업1", "TOTAL", CUR)
    sel = pwk_all[-1] if pwk_all else None
    if sel:
        render_insight(insight_product(sel))   # 상단 전체 요약(파란 박스)
        # 영업별 소계 요약(회색 줄)
        _subs = []
        for _ye in YEONG:
            _c = V("week", "product", "일평균거래액", _ye, "TOTAL", CUR, sel)
            _p = V("week", "product", "일평균거래액", _ye, "TOTAL", PREV, sel)
            if _c is not None:
                _subs.append(f"<b>{_ye}</b> {_c/1e6:,.0f}백만 {_pct(yoy(_c, _p))}")
        st.markdown(
            f"<div style='font-size:12px;color:#5b6472;background:#f7f9fc;border-left:3px solid #c5d3e8;"
            f"padding:6px 12px;border-radius:4px;margin:2px 0 8px'>기준: {week_pretty(sel)} · 거래액 일평균(백만) · "
            + _redneg("  |  ".join(_subs)) + "</div>", unsafe_allow_html=True)
        _tabs = st.tabs(YEONG)
        for _tab, _ye in zip(_tabs, YEONG):
            with _tab:
                render_gray_insight(insight_yeong(_ye, sel))   # 탭 인사이트는 회색 박스
                st.markdown(product_table_one(sel, _ye), unsafe_allow_html=True)

# ---- 종합 방향성 (BCG 스타일: 헤드라인 + 진단/실행/임팩트) ----
st.header("✅ 종합 방향성 및 전망", anchor="s6")
head_b, diag_b, now_b, nxt_b = final_direction(wk_all, cur_mo, cutoff)
render_bcg(head_b, diag_b, now_b, nxt_b)

with st.expander("ℹ️ 표 읽는 법 / 데이터"):
    st.markdown(
        f"- 각 표는 **구분 × [{CUR}년 | 전년비 | {PREV}년]** 구조로 엑셀 Summary 2.실적과 동일합니다.\n"
        "- **전년비**: 같은 월/주차 라벨을 올해 vs 전년으로 계산. 음수 <span style='color:#c0392b'>빨강 △</span>, 양수 검정.\n"
        "- 거래액=일평균거래액 **백만원**, 고객수·DAU=일평균, 유입률·CR=%. (유효회원수는 Summary처럼 숨김)\n"
        "- **월별은 깨끗한 일자별에서 집계**. 당월은 MTD(~집계일)이며 파란 테두리로 강조. 최신 주차도 동일 강조.\n"
        "- 일자별 차트는 **직전 2개 완료주 + 당주(집계일까지, 진행중)**, 전년은 **전년동요일**(날짜−364일) 비교.\n"
        "- ⚠️ 원본에 **2025년 9월말~10월** 손상 구간(유효회원수·DAU·거래액 ~2배)이 있어 자동 제외 → 해당 구간·전년비 '—'(2026은 정상).\n"
        "- 목표 대비 달성율(2-1)·행사별(2-6)은 원본 6종에 데이터가 없어 제외(목표 파일 주시면 추가).",
        unsafe_allow_html=True)

# ---- ✍️ 액션 방향 / 코멘트 (직접 작성) ----
COMMENT_FILE = "data/comment.md"
if "user_comment" not in st.session_state:
    try:
        with open(COMMENT_FILE, encoding="utf-8") as _f:
            st.session_state.user_comment = _f.read()
    except Exception:
        st.session_state.user_comment = ""

st.markdown("---")
with st.expander("✍️ 액션 방향 · 코멘트 (직접 작성)", expanded=bool(st.session_state.get("user_comment"))):
    st.caption("보고용 메모 — 데이터 갱신(재업로드)돼도 유지됩니다. '저장'하면 파일에도 남고, 아래로 내려받아 보관도 가능.")
    st.text_area("메모", key="user_comment", height=180,
                 placeholder="예) 금주 방향: DAU 회복(미방문 리텐션)에 집중, L+DAY 전환 극대화 …\n차주: SUMMER VACANCE 대비 …",
                 label_visibility="collapsed")
    _b1, _b2, _b3 = st.columns([1, 1, 4])
    if _b1.button("💾 저장"):
        try:
            with open(COMMENT_FILE, "w", encoding="utf-8") as _f:
                _f.write(st.session_state.user_comment or "")
            st.success("저장했습니다.")
        except Exception as _e:  # noqa
            st.warning(f"파일 저장 실패(세션엔 유지됨): {_e}")
    _b2.download_button("⬇️ 내려받기", st.session_state.get("user_comment", ""),
                        file_name="action_comment.md", mime="text/markdown")
